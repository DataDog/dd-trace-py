"""
openpyxl==3.0.10

https://pypi.org/project/openpyxl/
"""

import os
import tempfile

from flask import Blueprint
from flask import request

from .utils import ResultResponse


pkg_openpyxl = Blueprint("package_openpyxl", __name__)


@pkg_openpyxl.route("/openpyxl")
def pkg_openpyxl_view():
    import openpyxl

    response = ResultResponse(request.args.get("package_param"))

    try:
        param_value = request.args.get("package_param", "default-value")

        # Create a workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write the parameter value to the first cell
        ws["A1"] = param_value

        # Private directory per request: xdist workers share a cwd, so under a fixed name one
        # request's cleanup removes the file another request is still reading.
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = os.path.join(tmp_dir, "example.xlsx")
            wb.save(file_path)

            # Read back the value from the file to ensure it was written correctly
            wb_read = openpyxl.load_workbook(file_path)
            ws_read = wb_read.active
            read_value = ws_read["A1"].value

        result_output = f"Written value: {read_value}"

        response.result1 = result_output
    except Exception as e:
        response.result1 = f"Error: {str(e)}"

    return response.json()
